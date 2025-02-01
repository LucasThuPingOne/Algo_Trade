import pandas as pd
import matplotlib.pyplot as plt
from datetime import *
import os
from openpyxl import load_workbook, Workbook

# Charger la liste des tickers du secteur 'Information Technology'
sp500_url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
tables = pd.read_html(sp500_url)
sp500_table = tables[0]

sector = 'Utilities'

# Filtrer les entreprises dans le secteur 'Information Technology'
tech_sector = sp500_table[sp500_table['GICS Sector'] == sector]
tickers = tech_sector['Symbol'].tolist()

# Charger le fichier Excel
file_path = 'data/resultat_s&p500_trie.xlsx'  # Remplacez par le chemin de votre fichier Excel
xls = pd.ExcelFile(file_path)

# Dictionnaires pour stocker les rendements et prix de chaque entreprise
returns_dict = {}
price_dict = {}

# start_date = datetime(2023, 12, 31) - timedelta(days=2*365) 

# Spread de la data
t='5y'
# Créer le dossier 'data/Graph corr/2y/' s'il n'existe pas
output_dir = 'data/Graph corr/'+sector+'/'+t
os.makedirs(output_dir, exist_ok=True)

# Lire uniquement les feuilles correspondant aux tickers
for ticker in tickers:
    if ticker in xls.sheet_names:  # Vérifier si la feuille existe
        df = pd.read_excel(xls, sheet_name=ticker)
        
        # Convertir la colonne "Date" en datetime et la définir comme index
        df['date'] = pd.to_datetime(df['date'])
        # df = df[(df['date'] >= start_date)]
        df.set_index('date', inplace=True)
        
        # Remplacer la virgule par un point dans la colonne 'RET' si nécessaire
        df['RET'] = df['RET'].astype(str).str.replace(',', '.', regex=False)
        
        # Convertir la colonne 'RET' en type numérique
        df['RET'] = pd.to_numeric(df['RET'], errors='coerce')
        # Supprimer les lignes avec des dates en double pour éviter les erreurs de reindexation
        df = df[~df.index.duplicated(keep='first')]
        # Stocker les rendements et le prix
        returns_dict[ticker] = df['RET']
        price_dict[ticker] = df['PRC'] / df.loc[df.index[0], 'PRC']  # Normalisation

# Créer les DataFrames à partir des dictionnaires
returns_df = pd.DataFrame(returns_dict)
price_df = pd.DataFrame(price_dict)

# Calculer la matrice de corrélation
correlation_matrix = returns_df.corr()

# Trouver les paires d'entreprises avec une corrélation > 0.7 ou < -0.7
threshold = 0.7
correlation_pairs = []

# Boucle pour trouver les paires
for i in range(len(correlation_matrix.columns)):
    for j in range(i):
        if abs(correlation_matrix.iloc[i, j]) > threshold:
            correlation_pairs.append((correlation_matrix.columns[i], correlation_matrix.columns[j], correlation_matrix.iloc[i, j]))

# Convertir les paires en DataFrame pour un affichage facile
correlation_pairs_df = pd.DataFrame(correlation_pairs, columns=['Entreprise 1', 'Entreprise 2', 'Coefficient de Corrélation'])

# Afficher les paires
print("\nPaires d'entreprises avec corrélation > 0.7 ou < -0.7 :")
print(correlation_pairs_df)

# Tracer les graphes des paires
for pair in correlation_pairs:
    ticker1, ticker2, corr_value = pair

    plt.figure(figsize=(12, 6))
    plt.plot(price_df.index, price_df[ticker1], label=ticker1, alpha=0.7)
    plt.plot(price_df.index, price_df[ticker2], label=ticker2, alpha=0.7)

    plt.title(f'Prix des actions : {ticker1} et {ticker2}')
    plt.suptitle(f'Coefficient de corrélation : {corr_value:.2f}', fontsize=10)
    plt.xlabel('Date')
    plt.ylabel('Prix')
    plt.legend()
    plt.grid()
    
    # Sauvegarder le graphique dans le dossier spécifié
    plt.savefig(os.path.join(output_dir, f'{ticker1}_{ticker2}.png'), dpi=200)
    plt.close()  # Fermer la figure pour éviter d'utiliser trop de mémoire

# Enregistrer la matrice de corrélation dans un fichier Excel
output_file_path = 'data/Graph corr/'+sector+'/Correlation_matrix.xlsx'
os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
if not os.path.isfile(output_file_path):
    wb = Workbook()
    wb.save(output_file_path)
# correlation_matrix.to_excel(output_file_path, sheet_name='Correlation Matrix')

try:
        # Charger le fichier Excel s'il existe
        wb = load_workbook(output_file_path)
        # Supprimer la feuille si elle existe déjà pour éviter les doublons
        if t in wb.sheetnames:
            del wb[t]
        wb.save(output_file_path)

        # Ouvrir le fichier en mode append
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            correlation_matrix.to_excel(writer, sheet_name=t)  # Écrire les données dans une nouvelle feuille

except FileNotFoundError:
    # Si le fichier n'existe pas encore, le créer et y ajouter la première feuille
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        correlation_matrix.to_excel(writer, sheet_name=t)  # Créer une nouvelle feuille pour ce ticker

print(f"Matrice de corrélation enregistrée dans {output_file_path}")
