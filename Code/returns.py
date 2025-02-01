import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime
import os
from openpyxl import load_workbook 
from math import *

tickers = ['MSFT', 'AAPL']

end_date = datetime.today()
start_date = datetime(2020, 1, 1)

returns_df = pd.DataFrame()

#arbitraire (à voir pour fixer un seuil plus cohérent, qui dépend ou non de l'actif)
threshold = 0.03

output_folder = "data/"  #CHEMIN A MODIFIER SELON VOTRE EMPLACEMENT DANS VOS DOSSIERS
output_file = os.path.join(output_folder, 'Extreme_Daily_Returns2.xlsx')

for ticker in tickers:
    extreme_returns_df = pd.DataFrame()
    data = yf.download(ticker, start = start_date, end = end_date)
    #calcul du rendement jour par jour
    data['Day_return'] = data['Close']/data['Open'] - 1

    returns_df[ticker] = data['Day_return']
    #récupérer uniquement les rendements supérieurs en valeur absolue au seuil
    extreme_returns_df[ticker] = returns_df[ticker][abs(returns_df[ticker]) > threshold]
    try:
        wb = load_workbook(output_file)
        #supprimer la feuille si elle existe déjà
        if ticker in wb.sheetnames:
            del wb[ticker]
            
        wb.save(output_file)
    # Ouvrir le fichier en mode append
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        # Écrire le DataFrame dans la feuille correspondante, qui vient d'être (re)créée
            extreme_returns_df.to_excel(writer, sheet_name=ticker)

    except FileNotFoundError:
    # Si le fichier n'existe pas, le créer et ajouter la feuille
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            extreme_returns_df.to_excel(writer, sheet_name=ticker)







