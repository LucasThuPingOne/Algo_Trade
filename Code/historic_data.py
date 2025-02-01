import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook

# Récupération des composantes du S&P
sp500_url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
tables = pd.read_html(sp500_url)

# 1. Liste des composantes du S&P 500
tickers = tables[0]['Symbol'].tolist()  # Récupère la colonne 'Symbol'
tickers.append('^GSPC')  # Ajout du ticker pour l'indice S&P 500

# 2. Définir la période (1 an)
end_date = datetime.today()
start_date = end_date - timedelta(days=365)

# 3. Créer un dossier pour enregistrer les fichiers si nécessaire
output_folder = "data/"  # Assure-toi que le dossier existe ou crée-le
os.makedirs(output_folder, exist_ok=True)

# 4. Nom du fichier Excel de sortie
output_file = os.path.join(output_folder, 'SP500_Historic_Data.xlsx')

# 5. Télécharger les données pour chaque ticker
for ticker in tickers:
    print(f"Telechargement des donnees pour {ticker}...")
    
    # Télécharger les données avec yfinance
    data = yf.download(ticker, start=start_date, end=end_date)
    data['Ticker'] = ticker  # Ajouter une colonne pour identifier le ticker

    try:
        # Charger le fichier Excel s'il existe
        wb = load_workbook(output_file)
        # Supprimer la feuille si elle existe déjà pour éviter les doublons
        if ticker in wb.sheetnames:
            del wb[ticker]
        wb.save(output_file)

        # Ouvrir le fichier en mode append
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            data.to_excel(writer, sheet_name=ticker)  # Écrire les données dans une nouvelle feuille

    except FileNotFoundError:
        # Si le fichier n'existe pas encore, le créer et y ajouter la première feuille
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name=ticker)  # Créer une nouvelle feuille pour ce ticker

print(f"Donnees enregistrees dans le fichier : {output_file}")